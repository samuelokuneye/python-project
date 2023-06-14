import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
products_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, products_list.max_row + 1):
    products_list.cell(product_row, 4)
    supplier_name = products_list.cell(product_row, 4).value
    price = products_list.cell(product_row, 3).value
    inventory = products_list.cell(product_row, 2).value
    product_no = products_list.cell(product_row, 1).value
    inventory_price = products_list.cell(product_row, 5)

    #list of supplies and number of products supplied
    if supplier_name in product_per_supplier:
        current_no_product = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = current_no_product + 1
    else:
        print("adding new supplier")
        product_per_supplier[supplier_name] = 1

   #total value of inventory per supply
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

  #list of product with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_no)] = int(inventory)

 #writing the value of inventory in the spreadsheet
    inventory_price.value = inventory * price

print(products_under_10_inv)

print(f"The list of suppliers and number of products supplied are as follows:\n {product_per_supplier}")

print(f"The total value of inventory per supplier is:\n {total_value_per_supplier}")
inv_file.save("edited_inventory.xlsx")